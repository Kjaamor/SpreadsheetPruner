import os, openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

deletethisrowandbelow = 9

for filename in os.listdir('.'):
    if filename.endswith('.xlsx'):
        wb = openpyxl.load_workbook(filename)
        sheet = wb.active
        for row in range(deletethisrowandbelow, sheet.max_row + 1):
            sheet.delete_rows(deletethisrowandbelow)
        wb.save(filename)
        print('Done with ' + filename)